import datetime
import os.path
from datetime import date, datetime
import time
from threading import Timer
from openpyxl import *
from openpyxl.styles import *
from cal import Months
import sys
import os
from PyQt5 import QtWidgets
from PyQt5 import QtCore, QtGui
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import user_data


class Dash(QtWidgets.QWidget):

    switch_window = QtCore.pyqtSignal()

    def __init__(self):

        QtWidgets.QWidget.__init__(self)
        app.setStyle('Fusion')
        self.iconFile = 'atlogo.ico'

        self.status = " "

        data_file = "data.json"
        if not os.path.exists(data_file):
            name = self.getText()
            email = self.getText2()
            user_data.get_user_data(name, email)

        import initializer
        initializer.initialize()

        import mail
        if Clock.get_day(self) == 7:
            mail.send_mail()


        self.time_selected = 0
        user = user_data.extract_data()
        directory = user['directory']
        self.program_dir = directory
        self.directory_to = 'timesheets'
        self.file_name = Clock.get_year(self) + '_timebook.xlsx'

        self.threadpool = QThreadPool()

        # Init QSystemTrayIcon
        self.tray_icon = QSystemTrayIcon(QIcon(self.iconFile))
        self.tray_icon.show()

        # Tray menu
        show_action = QAction("Show", self)
        quit_action = QAction("Exit", self)
        hide_action = QAction("Hide", self)
        show_action.triggered.connect(self.show)
        hide_action.triggered.connect(self.hide)
        quit_action.triggered.connect(app.quit)

        tray_menu = QMenu()
        tray_menu.addAction(show_action)
        tray_menu.addAction(hide_action)
        tray_menu.addAction(quit_action)
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.show()

        # App window
        self.setGeometry(300, 300, 300, 180)
        self.setStyleSheet("background-color: White;")
        self.setWindowTitle("Auto Timesheet")
        self.setFixedSize(300, 180)
        self.center()

        # Widgets
        self.titlelabel = QtWidgets.QLabel(self)
        self.titlelabel.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.titlelabel.setText("")
        self.titlelabel.setPixmap(QtGui.QPixmap('CNRtitle.png'))
        self.titlelabel.setScaledContents(True)
        self.titlelabel.setObjectName("titlelabel")

        self.Label = QtWidgets.QLabel(self)
        self.Label.setText("Submission @ 16:00")
        self.Label.setAlignment(QtCore.Qt.AlignCenter)
        self.Label.move(85, 60)
        self.Label.setFont(QFont("calibri", 11))
        self.Label.adjustSize()


        self.submitButton = QtWidgets.QPushButton(self)
        self.submitButton.setText("Submit Now")
        self.submitButton.move(105, 100)
        self.submitButton.clicked.connect(self.submit_clicked)
        self.submitButton.setFont(QFont("Calibri", 11))
        self.submitButton.setStyleSheet("QPushButton::hover"
                                        "{"
                                        "background-color : lightgrey;"
                                        "}")

        self.timeButton = QtWidgets.QPushButton(self)
        self.timeButton.setText("Review")
        self.timeButton.move(107, 130)
        self.timeButton.clicked.connect(self.review_clicked)
        self.timeButton.setFont(QFont("Calibri", 11))

        self.Status = QtWidgets.QLabel(self)
        self.Status.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.Status.move(10, 165)
        self.Status.setText(str(self.status))
        self.Status.setObjectName("Promptlabel")
        self.Status.setFont(QFont("calibri", 6))
        self.Status.adjustSize()

        self.name = QtWidgets.QLabel(self)
        self.name.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.name.move(240, 165)
        self.name.setText("made by grizzly")
        self.name.setObjectName("Promptlabel")
        self.name.setFont(QFont("calibri", 6))
        self.name.adjustSize()

        self.wb = load_workbook(str(self.locate_timebook(self.program_dir)))

        self.Create()

        self.temp_wb = load_workbook(str(self.locate_template(self.program_dir)))

        self.active_sheet = self.wb[Clock.get_month(self)]
        self.check_timebook(self.program_dir)

        self.timer()

    def Create(self):
        if 'Sheet1' in self.wb.sheetnames:
            self.rename_Sheet1()
        else:
            pass
        if Clock.get_month(self) in self.wb.sheetnames:
            pass
        else:
            self.Status.setText('Sheet ' + Clock.get_month(self) + ' added in workbook')
            self.Status.adjustSize()
            self.wb.create_sheet(Clock.get_month(self))
            self.copy_from_template()
        self.Save()

    def rename_Sheet1(self):
        sheet = self.wb['Sheet1']
        sheet.title = 'Template'

    def copy_from_template(self):
        temp_sheet = self.wb['Template']
        active_sheet = self.wb[Clock.get_month(self)]
        for i in range(1, 7):
            for j in range(1, 4):
                active_sheet.cell(row=i, column=j).value = temp_sheet.cell(row=i, column=j).value

    def check_timebook(self, program_dir):
        directory_to = 'timesheets'
        file_name = Clock.get_year(self) + '_timebook.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        tb_exists = os.path.exists(program_dir + '/timesheets/' + file_name)
        if tb_exists == True:
            self.Status.setText("Timebook Exists")
            self.Status.adjustSize()
        else:
            self.temp_wb.save(os.path.join(directory_to, file_name))
            QtWidgets.QMessageBox.question(self, "Timebook Created",
                                           'Timebook for ' + Clock.get_year(self) + ' created',
                                           QtWidgets.QMessageBox.Ok)
            self.Status.setText("Timebook Created")
            self.Status.adjustSize()

    def locate_template(self, directory):
        program_dir = directory
        template_path = program_dir + '/templates/timesheet_template.xlsx'
        return template_path

    def locate_timebook(self, directory):
        program_dir = directory
        file_name = Clock.get_year(self) + '_timebook.xlsx'
        file_path = program_dir + '/timesheets/' + file_name
        return file_path

    def submit_clicked(self):
        self.Window_Switch()

    def review_clicked(self):
        filePath = self.locate_timebook(self.program_dir)
        os.system(f'start "excel" {filePath}')
        self.Status.setText("Opening Timebook...")
        self.Status.adjustSize()
        self.Status.setText(" ")

    def Window_Switch(self):
        time.sleep(0.15)
        self.switch_window.emit()

    def recurring_timer(self):
        self.time_selected += 1
        #self.l.setText("Counter: %d" % self.counter)
        self.Label.setText("Daily Submission at: %d" % self.timer)

    def closeEvent(self, event):
        event.ignore()
        self.hide()
        self.tray_icon.showMessage(
            "Auto Timesheet",
            "Application was minimized to Tray",
            QIcon(self.iconFile),
            2000
        )

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def getText(self):
        text, okPressed = QInputDialog.getText(self, "Enter Name", "Please enter your name and surname:", QLineEdit.Normal, "")
        if okPressed and text != '':
            return text

    def getText2(self):
        text, okPressed = QInputDialog.getText(self, "Enter email", "Please enter your email address:", QLineEdit.Normal, "")
        if okPressed and text != '':
            return text

    def timer(self):
        hour = 16
        minute = 00
        x = datetime.today()
        y = x.replace(day=x.day, hour=hour, minute=minute, second=0, microsecond=0)
        delta_t = y - x

        secs = delta_t.seconds + 1

        def Run():
            self.switch_window.emit()
            self.tray_icon.showMessage(
                "Auto Timesheet",
                "Time to enter daily submission.",
                QIcon(self.iconFile),
                2000
            )

        t = Timer(secs, Run)
        t.start()

    def Save(self):
        self.style()
        directory_to = 'timesheets'
        file_name = Clock.get_year(self) + '_timebook.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        self.wb.save(os.path.join(directory_to, file_name))
        self.Status.setText('File Saved')
        self.Status.adjustSize()


###-------------------------------------------------------------------###
class MainWindow(QtWidgets.QWidget):
    switch_window = QtCore.pyqtSignal()
    status = ""

    def __init__(self):
        global app
        QtWidgets.QWidget.__init__(self)
        app.setStyle('Fusion')

        data_file = "data.json"
        if not os.path.exists(data_file):
            name = self.getText()
            email = self.getText2()
            user_data.get_user_data(name, email)

        user = user_data.extract_data()
        directory = user['directory']
        self.program_dir = directory
        self.directory_to = 'timesheets'
        self.file_name = Clock.get_year(self) + '_timebook.xlsx'
        #self.charCount = 0

        # App window
        self.setGeometry(300, 300, 300, 180)
        self.setStyleSheet("background-color: White;")
        self.setWindowTitle("Auto Timesheet")
        self.setFixedSize(300, 180)
        self.center()

        # Widgets
        self.Promptlabel = QtWidgets.QLabel(self)
        self.Promptlabel.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.Promptlabel.move(70, 20)
        self.Promptlabel.setText("Today's work description:")
        self.Promptlabel.setObjectName("Promptlabel")
        self.Promptlabel.setFont(QFont("calibri", 13))
        self.Promptlabel.adjustSize()

        self.line = QLineEdit(self)
        self.line.move(25, 50)
        self.line.resize(250, 28)

        self.submitButton = QtWidgets.QPushButton(self)
        self.submitButton.setText("Submit")
        self.submitButton.move(107, 100)
        self.submitButton.clicked.connect(self.submit_clicked)
        self.submitButton.setFont(QFont("Calibri", 11))
        self.submitButton.setStyleSheet("QPushButton::hover"
                                      "{"
                                      "background-color : lightgrey;"
                                      "}")

        self.timeButton = QtWidgets.QPushButton(self)
        self.timeButton.setText("Back")
        self.timeButton.move(107, 130)
        self.timeButton.clicked.connect(self.back_clicked)
        self.timeButton.setFont(QFont("Calibri", 11))
        self.timeButton.setStyleSheet("QPushButton::hover"
                                        "{"
                                        "background-color : lightgrey;"
                                        "}")

        self.Status = QtWidgets.QLabel(self)
        self.Status.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.Status.move(10, 165)
        self.Status.setText(str(self.status))
        self.Status.setObjectName("Promptlabel")
        self.Status.setFont(QFont("calibri", 6))
        self.Status.adjustSize()

        self.name = QtWidgets.QLabel(self)
        self.name.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.name.move(240, 165)
        self.name.setText("made by grizzly")
        self.name.setObjectName("Promptlabel")
        self.name.setFont(QFont("calibri", 6))
        self.name.adjustSize()

        self.temp_wb = load_workbook(str(self.locate_template(self.program_dir)))
        self.wb = load_workbook(str(self.locate_timebook(self.program_dir)))
        self.Create()
        self.active_sheet = self.wb[Clock.get_month(self)]

    def Create(self):
        if 'Sheet1' in self.wb.sheetnames:
            self.rename_Sheet1()
        else:
            pass
        if Clock.get_month(self) in self.wb.sheetnames:
            pass
        else:
            self.Status.setText('Sheet ' + Clock.get_month(self) + ' added in workbook')
            self.Status.adjustSize()
            self.wb.create_sheet(Clock.get_month(self))
            self.copy_from_template()

    def rename_Sheet1(self):
        sheet = self.wb['Sheet1']
        sheet.title = 'Template'

    def copy_from_template(self):
        temp_sheet = self.wb['Template']
        active_sheet = self.wb[Clock.get_month(self)]
        for i in range(1, 7):
            for j in range(1, 4):
                active_sheet.cell(row=i, column=j).value = temp_sheet.cell(row=i, column=j).value

    def day_Entry(self):
        n = 40
        lineCount = 1
        desc_value = self.line_value()
        desc_item = desc_value.split(';')
        for idx in desc_item:
            desc_line = [idx[i:i + n] for i in range(0, len(idx), n)]
            for i in desc_line:
                 if lineCount == 1:
                     day_list = [(str(Clock.get_day(self)), i, '*')]
                     for d in day_list:
                         self.active_sheet.append(d)
                     lineCount += 1
                 else:
                     day_list = [('', i, '*')]
                     for d in day_list:
                         self.active_sheet.append(d)
                     lineCount += 1
        self.line.setText("")
        self.line.setPlaceholderText(f'Description for {Clock.get_day(self)} {Clock.get_month(self)} submitted.')



    def curr_week(self):
        week = Clock.get_week_of_month(self)
        curr_week = 'Week' + str(week)
        return str(curr_week)

    def month_Entries(self):
        current_week = self.curr_week()
        if date.today().weekday() == 0:
            if self.active_sheet.cell(row=self.active_sheet.max_row,column=1).value == Clock.get_day(self):
                self.line.setText("")
                self.line.setPlaceholderText('Daily Entry Satisfied')
            elif self.active_sheet.cell(row=self.active_sheet.max_row, column=1).value == '':
                self.line.setText("")
                self.line.setPlaceholderText('Daily Entry Satisfied')
            elif self.active_sheet.cell(row=self.active_sheet.max_row,column=1).value == 'PROJECT':
                self.active_sheet.cell(row=self.active_sheet.max_row+1,column=1).value = current_week
                self.day_Entry()
            else:
                if self.active_sheet.cell(row=self.active_sheet.max_row,column=1).value == current_week:
                    self.day_Entry()
                else:
                    self.active_sheet.cell(row=self.active_sheet.max_row+1,column=1).value = current_week
                    self.day_Entry()
        elif not date.today().weekday() == 0 or 5 or 6:
            if self.active_sheet.cell(row=self.active_sheet.max_row,column=1).value == Clock.get_day(self):
                self.line.setText("")
                self.line.setPlaceholderText('Daily Entry Satisfied')
            elif self.active_sheet.cell(row=self.active_sheet.max_row,column=1).value == 'PROJECT':
                self.active_sheet.cell(row=self.active_sheet.max_row+1,column=1).value = current_week
                self.day_Entry()
            else:
                self.day_Entry()


    def line_value(self):
        return self.line.text()


    def submit_clicked(self):
        self.month_Entries()
        self.Save()


    def back_clicked(self):
        self.switch_window.emit()


    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def locate_template(self, directory):
        program_dir = directory
        template_path = program_dir + '/templates/timesheet_template.xlsx'
        return template_path


    def locate_timebook(self, directory):
        program_dir = directory
        file_name = Clock.get_year(self) + '_timebook.xlsx'
        file_path = program_dir + '/timesheets/' + file_name
        return file_path


    def Column_sizes(self):
        active_sheet = self.wb[Clock.get_month(self)]
        active_sheet.column_dimensions['A'].width = 17
        active_sheet.column_dimensions['B'].width = 48
        active_sheet.column_dimensions['C'].width = 17



    def Borders(self):
        active_sheet = self.wb[Clock.get_month(self)]
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

        for i in range(1, 4):
            active_sheet.cell(row=6, column=i).border = Border(top=thin, left=thin, right=thin, bottom=double)

        for i in range(7, active_sheet.max_row + 1):
            for j in range(1, 4):
                active_sheet.cell(row=i, column=j).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def Headers(self):
        active_sheet = self.wb[Clock.get_month(self)]
        medium_font = Font(name='Calibri', bold=True, size=16)
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        large_font = Font(name='Calibri', bold=True, size=18)

        # ---Timesheet---#
        active_sheet['A1'].value = 'TIMESHEET'
        active_sheet['A1'].font = large_font

        active_sheet['B1'].value = 'cnr'
        active_sheet['C1'].value = '.Architects'
        active_sheet['B1'].font = Font(name='Bauhaus 93', size=16, bold=True)
        active_sheet['C1'].font = Font(name='New Times Roman', size=16, bold=False)
        active_sheet['B1'].alignment = Alignment(horizontal='right')

        for i in range(1, 4):
            active_sheet.cell(row=6, column=i).font = medium_font
            active_sheet.cell(row=6, column=i).alignment = center_align

    def style(self):
        self.Column_sizes()
        self.Borders()
        self.Headers()

    def getText(self):
        text, okPressed = QInputDialog.getText(self, "Enter Name", "Please enter your name and surname:", QLineEdit.Normal, "")
        if okPressed and text != '':
            return text

    def getText2(self):
        text, okPressed = QInputDialog.getText(self, "Enter email", "Please enter your email address:", QLineEdit.Normal, "")
        if okPressed and text != '':
            return text

    def EnterCredentials(self):
        active_sheet = self.wb[Clock.get_month(self)]
        current_month = str(Clock.get_month(self))
        current_year = str(Clock.get_year(self))

        if active_sheet['B3'].value == 'xxx':
            active_sheet['B3'].value = current_month + ' ' + current_year
            self.Save()
        if active_sheet['B4'].value == 'xxx':
            user = user_data.extract_data()
            new_user = user['user']
            active_sheet['B4'].value = new_user['name']
            self.Save()
        else:
            pass


    def Save(self):
        self.style()
        directory_to = 'timesheets'
        file_name = Clock.get_year(self) + '_timebook.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        self.wb.save(os.path.join(directory_to, file_name))
        self.Status.setText('File Saved')
        self.Status.adjustSize()


class Clock:
    def get_day(self):
        today = date.today()
        day_num = today.day
        return str(day_num)

    def get_month(self):
        today = date.today()
        month_num = today.month
        month_name = Months[str(month_num)]
        return str(month_name)

    def get_year(self):
        today = date.today()
        year_num = today.year
        return str(year_num)

    def get_week_of_month(self):
        day_of_month = datetime.now().day
        week_number = (day_of_month - 1) // 7 + 1
        return week_number


##----------------------------------------------------------------------##
class Controller:

    def __init__(self):
        self.dash = Dash()
        self.submit = MainWindow()

    def show_dash(self):
        self.dash.switch_window.connect(self.show_submit)
        self.submit.hide()
        self.dash.center()
        self.dash.Status.setText('')
        self.dash.show()

    def show_submit(self):
        self.submit.switch_window.connect(self.show_dash)
        self.dash.hide()
        self.submit.center()

        if self.submit.line.text() != f'Description for {Clock.get_day(self)} {Clock.get_month(self)} submitted.':
            self.submit.line.setPlaceholderText('')

        if self.submit.Status.text() == "File Saved":
            self.submit.Status.setText("")

        self.submit.show()


if __name__ == '__main__':
    iconFile = 'atlogo.ico'
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QtWidgets.QApplication(sys.argv)    
    app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    app.setQuitOnLastWindowClosed(False)
    controller = Controller()
    app.setWindowIcon(QIcon(iconFile))
    controller.show_dash()
    sys.exit(app.exec_())


