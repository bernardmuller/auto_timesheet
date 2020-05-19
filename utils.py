from openpyxl import *
import os, os.path
from  datetime import  date
from cal import Months


## there should always be a template file in the templates directory ##
def locate_template(directory):
    program_dir = directory    
    template_path = program_dir + '/templates/timesheet_template.xlsx'
    return template_path    


def locate_timebook(directory):
    program_dir = directory
    file_name = str(get_year()) + '_timeheets.xlsx'
    file_path = program_dir + '/timesheets/' + file_name
    return file_path


def get_month():
    today = date.today()
    month_num = today.month
    month_name = Months[str(month_num)]
    return month_name


def get_year():
    today = date.today()
    year_num = today.year    
    return year_num


def create_timebook(program_dir):
    directory_to = 'timesheets'
    file_name = str(get_year()) + '_timeheets.xlsx'
    if not os.path.isdir(directory_to):
        os.makedirs(directory_to)
    tb_exists = os.path.exists(program_dir + '/timesheets/' + file_name )
    if tb_exists == True:
        ## Prompt if file already exists , override? ##
        print('Timebook for ' + str(get_year()) + ' already exists')
        prompt= input('Override? Y/N :')
        if prompt == 'Y' or 'y':
            temp_wb.save(os.path.join(directory_to, file_name))
            print('Timebook for ' + str(get_year()) + ' created')
        else:
            pass     
    else:
        temp_wb.save(os.path.join(directory_to, file_name)) 
        print('Timebook for ' + str(get_year()) + 'created')  

        
# def check_timeheet():
#     if str(get_month()) in wb.sheetnames:
#         pass
#     else:
#         wb.add_sheet(book[str(get(month))])        
#         pass

   

if __name__ == "__main__": 
    ## user required to insert path to program ##
    program_dir = 'C:/Users/Bernard/Dropbox/Personal/python/auto_timesheet'
    
    ## template workbook object ##
    temp_wb = load_workbook(str(locate_template(program_dir)))
    ## current year workbook object ##
    wb = load_workbook(str(locate_timebook(program_dir)))

    ## checks if current years timebook exists, if not creates new ##
    create_timebook(program_dir)  
         
    
