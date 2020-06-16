import datetime
import os
import os.path
from datetime import date
import time
from dir import Setup_Dir

from openpyxl import *
from openpyxl.styles import *

from cal import Months

import sys
import os
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QAction, QApplication, QLabel, QMainWindow, QMenu, QStyle, QSystemTrayIcon, QDesktopWidget, \
    QLineEdit, QStatusBar
from PyQt5 import QtCore, QtGui
from PyQt5.QtGui import *


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


class MainWindow(QtWidgets.QWidget):
    switch_window = QtCore.pyqtSignal()
    iconFile = resource_path('icon/CNRlogo.ico')

    def __init__(self):
        global app
        QtWidgets.QWidget.__init__(self)
        app.setStyle('Fusion')

        # App window
        self.setGeometry(300, 300, 300, 180)
        self.setStyleSheet("background-color: White;")
        self.setWindowTitle("Auto Timesheet")
        self.setFixedSize(300, 180)
        self.center()

        # Widgets
        self.Promptlabel = QtWidgets.QLabel(self)
        self.Promptlabel.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.Promptlabel.move(70, 20)
        self.Promptlabel.setText("Today's work description:")
        self.Promptlabel.setObjectName("Promptlabel")
        self.Promptlabel.setFont(QFont("calibri", 13))
        self.Promptlabel.adjustSize()

        self.line = QLineEdit(self)
        self.line.move(25, 50)
        self.line.resize(250, 28)

        self.submitButton = QtWidgets.QPushButton(self)
        self.submitButton.setText("Submit")
        self.submitButton.move(107, 100)
        self.submitButton.clicked.connect(self.submit_clicked)
        self.submitButton.setFont(QFont("Calibri", 11))
        self.submitButton.setStyleSheet("QPushButton::hover"
                                      "{"
                                      "background-color : lightgrey;"
                                      "}")

        self.timeButton = QtWidgets.QPushButton(self)
        self.timeButton.setText("Back")
        self.timeButton.move(107, 130)
        self.timeButton.clicked.connect(self.back_clicked)
        self.timeButton.setFont(QFont("Calibri", 11))
        self.timeButton.setStyleSheet("QPushButton::hover"
                                        "{"
                                        "background-color : lightgrey;"
                                        "}")


        self.name = QtWidgets.QLabel(self)
        self.name.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.name.move(240, 165)
        self.name.setText("made by grizzly")
        self.name.setObjectName("Promptlabel")
        self.name.setFont(QFont("calibri", 6))
        self.name.adjustSize()

        # self.statusbar = QtWidgets.QStatusBar(MainWindow)
        # self.statusbar.setObjectName("statusbar")
        # MainWindow.setStatusBar(self.statusbar)
        # self.statusBar().showMessage('Message in statusbar.')

    def submit_clicked(self):
        self.line.setPlaceholderText("Coming soon...")
        pass


    def back_clicked(self):
        self.switch_window.emit()
        pass

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


class WindowTwo(QtWidgets.QWidget):

    def __init__(self, text):
        QtWidgets.QWidget.__init__(self)
        self.setWindowTitle('Window Two')

        layout = QtWidgets.QGridLayout()

        self.label = QtWidgets.QLabel(text)
        layout.addWidget(self.label)

        self.button = QtWidgets.QPushButton('Close')
        self.button.clicked.connect(self.close)

        layout.addWidget(self.button)

        self.setLayout(layout)


class Dash(QtWidgets.QWidget):

    switch_window = QtCore.pyqtSignal()
    iconFile = resource_path('icon/CNRlogo.ico')

    def __init__(self):
        global app
        QtWidgets.QWidget.__init__(self)
        app.setStyle('Fusion')
        self.iconFile = resource_path('icon/CNRlogo.ico')

        # Init QSystemTrayIcon
        self.tray_icon = QSystemTrayIcon(QIcon(self.iconFile))
        self.tray_icon.show()

        # Tray menu
        show_action = QAction("Show", self)
        quit_action = QAction("Exit", self)
        hide_action = QAction("Hide", self)
        show_action.triggered.connect(self.show)
        hide_action.triggered.connect(self.hide)
        quit_action.triggered.connect(app.quit)

        tray_menu = QMenu()
        tray_menu.addAction(show_action)
        tray_menu.addAction(hide_action)
        tray_menu.addAction(quit_action)
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.show()

        # App window
        self.setGeometry(300, 300, 300, 180)
        self.setStyleSheet("background-color: White;")
        self.setWindowTitle("Auto Timesheet")
        self.setFixedSize(300, 180)
        self.center()

        # Widgets

        self.titlelabel = QtWidgets.QLabel(self)
        self.titlelabel.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.titlelabel.setText("")
        self.titlelabel.setPixmap(QtGui.QPixmap(resource_path('icon/CNRtitle.png')))
        self.titlelabel.setScaledContents(True)
        self.titlelabel.setObjectName("titlelabel")

        self.Label = QtWidgets.QLabel(self)
        self.Label.move(70, 60)
        self.Label.setAlignment(QtCore.Qt.AlignCenter)
        self.Label.setText("Next Submission at: (time) ")
        self.Label.setFont(QFont("calibri", 11))
        self.Label.adjustSize()

        self.submitButton = QtWidgets.QPushButton(self)
        self.submitButton.setText("Submit Now")
        self.submitButton.move(105, 100)
        self.submitButton.clicked.connect(self.submit_clicked)
        self.submitButton.setFont(QFont("Calibri", 11))
        self.submitButton.setStyleSheet("QPushButton::hover"
                                        "{"
                                        "background-color : lightgrey;"
                                        "}")

        # self.timeButton = QtWidgets.QPushButton(self)
        # self.timeButton.setText("Set Time")
        # self.timeButton.move(107.5, 130)
        # self.timeButton.clicked.connect(self.time_clicked)
        # self.timeButton.setFont(QFont("Calibri", 11))


        self.name = QtWidgets.QLabel(self)
        self.name.setGeometry(QtCore.QRect(50, 10, 200, 40))
        self.name.move(240, 165)
        self.name.setText("made by grizzly")
        self.name.setObjectName("Promptlabel")
        self.name.setFont(QFont("calibri", 6))
        self.name.adjustSize()

        # self.statusbar = QtWidgets.QStatusBar()
        # self.statusbar.setObjectName("statusbar")
        # MainWindow.setStatusBar(self.statusbar)
        # self.statusBar().showMessage('Message in statusbar.')

    def submit_clicked(self):
        self.switch_window.emit()

    def closeEvent(self, event):
        event.ignore()
        self.hide()
        self.tray_icon.showMessage(
            "Auto Timesheet",
            "Application was minimized to Tray",
            QIcon(self.iconFile),
            2000
        )

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


class Controller:

    def __init__(self):
        self.dash = Dash()
        self.submit = MainWindow()


    def show_dash(self):
        #self.dash = Dash()
        self.dash.switch_window.connect(self.show_submit)
        self.submit.hide()
        self.dash.show()

    def show_submit(self):
        #self.submit = MainWindow()
        self.submit.switch_window.connect(self.show_dash)
        self.dash.hide()
        self.submit.show()

    def show_window_two(self, text):
        self.window_two = WindowTwo(text)
        self.window.close()
        self.window_two.show()


class ProgramSetup:

    def __init__(self):
        pass

class Clock:

    def get_day(self):
        today = date.today()
        day_num = today.day
        # month_name = Months[str(month_num)]
        return str(day_num)

    def get_month(self):
        today = date.today()
        month_num = today.month
        month_name = Months[str(month_num)]
        return str(month_name)

    def get_year(self):
        today = date.today()
        year_num = today.year
        return str(year_num)

    def get_week_of_month(self):
        day_of_month = datetime.datetime.now().day
        week_number = (day_of_month - 1) // 7 + 1
        return week_number


class TimebookSetup:

    def Save(self):
        directory_to = 'timesheets'
        file_name = Clock.get_year() + '_timeheets.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        wb.save(os.path.join(directory_to, file_name))
        print('File Saved')

    ## there should always be a template file in the templates directory ##
    def locate_template(self, directory):
        program_dir = directory
        template_path = program_dir + '/templates/timesheet_template.xlsx'
        return template_path

    def locate_timebook(self, directory):
        program_dir = directory
        file_name = Clock.get_year() + '_timeheets.xlsx'
        file_path = program_dir + '/timesheets/' + file_name
        return file_path

    def Create(self, program_dir):
        directory_to = 'timesheets'
        file_name = Clock.get_year() + '_timeheets.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        tb_exists = os.path.exists(program_dir + '/timesheets/' + file_name)
        if tb_exists == True:
            ## Prompt if file already exists , override? ##
            print('Timebook for ' + Clock.get_year() + ' already exists')
            prompt = input('Override? Y/N :')
            if prompt in ['Y', 'y']:
                prompt2 = input('Are you sure? Y/N :')
                if prompt2 in ['Y', 'y']:
                    temp_wb.save(os.path.join(directory_to, file_name))
                    print('Timebook for ' + Clock.get_year() + ' created')
                else:
                    pass
            else:
                pass
        else:
            temp_wb.save(os.path.join(directory_to, file_name))
            print('Timebook for ' + Clock.get_year() + ' created')


class TimesheetSetup:

    def Create(self):
        if 'Sheet1' in wb.sheetnames:
            self.rename_Sheet1()
        else:
            pass
        if Clock.get_month() in wb.sheetnames:
            pass
        else:
            print('Sheet ' + Clock.get_month() + ' added in workbook')
            wb.create_sheet(Clock.get_month())
            self.copy_from_template()
            self.EnterCredentials()

    def rename_Sheet1(self):
        sheet = wb['Sheet1']
        sheet.title = 'Template'

    def copy_from_template(self):
        temp_sheet = wb['Template']
        active_sheet = wb[Clock.get_month()]
        for i in range(1, 7):
            for j in range(1, 4):
                active_sheet.cell(row=i, column=j).value = temp_sheet.cell(row=i, column=j).value

    def EnterCredentials(self):
        active_sheet = wb[Clock.get_month()]
        current_month = str(Clock.get_month())
        current_year = str(Clock.get_year())
        if active_sheet['B3'].value == 'xxx':
            active_sheet['B3'].value = current_month + ' ' + current_year
        if active_sheet['B4'].value == 'xxx':
            new_user = input('Please enter your name: ')
            active_sheet['B4'].value = new_user
            print('Username saved.')
        else:
            pass


class Editor:

    def day_Entry(self):

        desc_value = input('Today\'s work description: ')
        day_list = [(str(Clock.get_day()), desc_value, '*')]
        for d in day_list:
            active_sheet.append(d)
        print(f'Description for {Clock.get_day()} {Clock.get_month()} submitted.')

    def curr_week(self):
        week = Clock.get_week_of_month()
        curr_week = 'Week' + str(week)
        return str(curr_week)

        ## following appends work descriptions to excel file

    def month_Entries(self):
        current_week = self.curr_week()
        if date.today().weekday() == 0:
            if active_sheet.cell(row=active_sheet.max_row, column=1).value == Clock.get_day():
                print('Daily Entry Satisfied')
            elif active_sheet.cell(row=active_sheet.max_row, column=1).value == 'PROJECT':
                active_sheet.cell(row=active_sheet.max_row + 1, column=1).value = current_week
                self.day_Entry()
            else:
                if active_sheet.cell(row=active_sheet.max_row, column=1).value == current_week:
                    self.day_Entry()
                else:
                    active_sheet.cell(row=active_sheet.max_row + 1, column=1).value = current_week
                    self.day_Entry()
        elif not date.today().weekday() == 0 or 5 or 6:
            if active_sheet.cell(row=active_sheet.max_row, column=1).value == Clock.get_day():
                print('Daily Entry Satisfied')
            elif active_sheet.cell(row=active_sheet.max_row, column=1).value == 'PROJECT':
                active_sheet.cell(row=active_sheet.max_row + 1, column=1).value = current_week
                self.day_Entry()
            else:
                self.day_Entry()


class Styler:

    def Column_sizes(self):
        active_sheet = wb[Clock.get_month()]
        active_sheet.column_dimensions['A'].width = 17
        active_sheet.column_dimensions['B'].width = 48
        active_sheet.column_dimensions['C'].width = 17

    def Borders(self):
        active_sheet = wb[Clock.get_month()]
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

        for i in range(1, 4):
            active_sheet.cell(row=6, column=i).border = Border(top=thin, left=thin, right=thin, bottom=double)

        for i in range(7, active_sheet.max_row + 1):
            for j in range(1, 4):
                active_sheet.cell(row=i, column=j).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def Headers(self):
        active_sheet = wb[Clock.get_month()]
        medium_font = Font(name='Calibri', bold=True, size=16)
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        large_font = Font(name='Calibri', bold=True, size=18)

        # ---Timesheet---#
        active_sheet['A1'].value = 'TIMESHEET'
        active_sheet['A1'].font = large_font

        active_sheet['B1'].value = 'CNR'
        active_sheet['C1'].value = '.Architects'
        active_sheet['B1'].font = Font(name='Bauhaus 93', size=16, bold=True)
        active_sheet['C1'].font = Font(name='New Times Roman', size=16, bold=False)
        active_sheet['B1'].alignment = Alignment(horizontal='right')

        for i in range(1, 4):
            active_sheet.cell(row=6, column=i).font = medium_font
            active_sheet.cell(row=6, column=i).alignment = center_align

    def style(self):
        self.Column_sizes()
        self.Borders()
        self.Headers()


def Run():
    global ProgramSetup
    global program_dir
    global Clock
    global TimebookSetup
    global TimesheetSetup
    global Editor
    global Style
    global temp_wb
    global wb
    global active_sheet

    TimebookSetup.Create(program_dir)
    TimesheetSetup.Create()
    Editor.month_Entries()
    Style.style()
    TimebookSetup.Save()
    time.sleep(5)

def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    controller = Controller()
    app.setWindowIcon(QIcon(iconFile))
    controller.show_dash()
    sys.exit(app.exec_())


app = QtWidgets.QApplication(sys.argv)
iconFile = resource_path('icon/CNRlogo.ico')
ProgramSetup = ProgramSetup()
program_dir = Setup_Dir()
Clock = Clock()
TimebookSetup = TimebookSetup()
TimesheetSetup = TimesheetSetup()
Editor = Editor()
Style = Styler()
temp_wb = load_workbook(str(TimebookSetup.locate_template(program_dir)))
wb = load_workbook(str(TimebookSetup.locate_timebook(program_dir)))
active_sheet = wb[Clock.get_month()]


if __name__ == '__main__':
    #Run()
    main()